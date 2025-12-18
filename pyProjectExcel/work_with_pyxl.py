import openpyxl as op
import pprint as pp
from libs import *

filename = 'Бланк заказа.xlsx'

wb = op.load_workbook(filename, data_only=True)
sheet1 = wb.active

subcat_dict = {}

max_rows = sheet1.max_row

for i in range(7, max_rows):
    sku = sheet1.cell(row=i, column=2).value
    subcat = sheet1.cell(row=i, column=10).value

    if not sku:
        continue

    if subcat not in subcat_dict:
        subcat_dict[subcat] = [sku]
    else:
        subcat_dict[subcat].append(sku)

with open('report_subcat_sku.txt', 'w') as myfile:

    for key, value in subcat_dict.items():
        values = ', '.join(value)
        string_to_write = key + ': ' + values + '\n'
        myfile.write(string_to_write)

mult_my_list